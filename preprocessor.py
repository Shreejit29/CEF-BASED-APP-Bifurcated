# preprocessor.py
import io
import zipfile
import openpyxl
import streamlit as st

# Canonical headers expected downstream
NEW_HEADINGS = [
    "Cycle_Number",       # canonical instead of "Cycle number"
    "Time",
    "Date",
    "Voltage (mV)",
    "Current (mA)",
    "Capacity (mAh)",
    "Energy (mWh)"
]

# 1-based indices to delete (after keeping the 2nd sheet)
COLUMNS_TO_DELETE = [3, 9, 10, 11, 12, 13]

def _process_workbook_bytes(file_bytes: bytes, filename: str):
    issues = []
    try:
        wb = openpyxl.load_workbook(io.BytesIO(file_bytes))
    except Exception as e:
        issues.append(f"{filename}: load error: {e}")
        return None, f"modified_{filename}", issues

    sheet_names = wb.sheetnames
    if len(sheet_names) < 2:
        issues.append(f"{filename}: not enough sheets; skipped.")
        return None, f"modified_{filename}", issues

    target_name = sheet_names[1]
    ws = wb[target_name]

    # Remove all other sheets
    for sn in list(sheet_names):
        if sn != target_name:
            wb.remove(wb[sn])

    # Delete specified columns (descending preserves positions)
    for col in sorted(COLUMNS_TO_DELETE, reverse=True):
        try:
            ws.delete_cols(col)
        except Exception as e:
            issues.append(f"{filename}: delete col {col} failed: {e}")

    # Overwrite header row with canonical names
    for idx, heading in enumerate(NEW_HEADINGS, start=1):
        ws.cell(row=1, column=idx, value=heading)

    # Rename sheet to base filename
    base = filename.rsplit(".", 1)[0]
    try:
        ws.title = base[:31]  # Excel sheet title limit
    except Exception:
        ws.title = "Sheet1"

    # Save to bytes
    out = io.BytesIO()
    wb.save(out)
    out.seek(0)
    return out.read(), f"modified_{filename}", issues

def render_preprocessor():
    st.subheader("Pre-Processor")
    st.caption("Keeps only the 2nd sheet, deletes columns [3,9,10,11,12,13], rewrites headers to canonical names, and renames the sheet to the filename.")
    uploaded = st.file_uploader("Upload Excel files", type=["xlsx"], accept_multiple_files=True)

    auto_continue = st.checkbox("Send pre-processed files to Raw Processing automatically", value=False)

    if st.button("Run Pre-Processor") and uploaded:
        modified = []
        report = []
        for f in uploaded:
            try:
                content = f.read()
                f.seek(0)
                out_bytes, out_name, issues = _process_workbook_bytes(content, f.name)
                if issues:
                    report.extend(issues)
                if out_bytes is not None:
                    modified.append((out_name, out_bytes))
            except Exception as e:
                report.append(f"{f.name}: unexpected error: {e}")

        if not modified:
            st.warning("No files produced. Check the report below.")
        else:
            # Optional download
            with st.expander("Download pre-processed files (optional)", expanded=False):
                if len(modified) == 1:
                    name, data = modified[0]
                    st.download_button(
                        f"Download {name}", data=data, file_name=name,
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        use_container_width=True,
                    )
                else:
                    buf = io.BytesIO()
                    with zipfile.ZipFile(buf, "w", compression=zipfile.ZIP_DEFLATED) as zf:
                        for name, data in modified:
                            zf.writestr(name, data)
                    buf.seek(0)
                    st.download_button(
                        "Download all (ZIP)", data=buf, file_name="modified_excels.zip",
                        mime="application/zip", use_container_width=True,
                    )

            # Auto-continue: stash into session for the Raw page to consume
            if auto_continue:
                class TempUpload(io.BytesIO):
                    def __init__(self, data, name):
                        super().__init__(data)
                        self.name = name
                st.session_state["_forwarded_preproc_files"] = [TempUpload(b, n) for n, b in modified]
                st.success("Pre-processed files are ready for Raw Processing from session.")

        if report:
            st.markdown("#### Report")
            for line in report:
                st.write("-", line)
