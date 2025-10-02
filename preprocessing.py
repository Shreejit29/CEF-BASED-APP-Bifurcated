import io
import zipfile
import openpyxl
import streamlit as st
import tempfile

# Canonical headers expected downstream
NEW_HEADINGS = [
    "Cycle_Number",
    "Time",
    "Date",
    "Voltage (mV)",
    "Current (mA)",
    "Capacity (mAh)",
    "Energy (mWh)"
]

# 1-based indices to delete on the kept sheet (2nd sheet)
COLUMNS_TO_DELETE = [3, 9, 10, 11, 12, 13]

def _process_workbook_bytes(file_bytes: bytes, filename: str):
    issues = []
    try:
        wb = openpyxl.load_workbook(io.BytesIO(file_bytes))
    except Exception as e:
        issues.append(f"{filename}: load error: {e}")
        return None, f"modified_{filename}", issues

    sheet_names = wb.sheetnames
    if len(sheet_names) < 2:
        issues.append(f"{filename}: not enough sheets; skipped.")
        return None, f"modified_{filename}", issues

    target_name = sheet_names[1]
    ws = wb[target_name]

    # Remove other sheets
    for sn in list(sheet_names):
        if sn != target_name:
            wb.remove(wb[sn])

    # Delete columns in descending order
    for col in sorted(COLUMNS_TO_DELETE, reverse=True):
        try:
            ws.delete_cols(col)
        except Exception as e:
            issues.append(f"{filename}: delete col {col} failed: {e}")

    # Canonical header row
    for idx, heading in enumerate(NEW_HEADINGS, start=1):
        ws.cell(row=1, column=idx, value=heading)

    # Rename sheet to filename base (Excel <=31 chars)
    base = filename.rsplit(".", 1)[0]
    try:
        ws.title = base[:31]
    except Exception:
        ws.title = "Sheet1"

    # Serialize
    out = io.BytesIO()
    wb.save(out)
    out.seek(0)
    return out.read(), f"modified_{filename}", issues

def render_preprocessor():
    st.subheader("Pre-Processor")
    st.caption("Keeps only the 2nd sheet, deletes columns [3,9,10,11,12,13], writes canonical headers, renames sheet to filename, and forwards output to Raw processing.")

    uploaded = st.file_uploader("Upload Excel files", type=["xlsx"], accept_multiple_files=True)
    auto_continue = st.checkbox("Send output to Raw Processing automatically", value=True)

    if st.button("Run Pre-Processor") and uploaded:
        modified, report = [], []
        for f in uploaded:
            try:
                content = f.read(); f.seek(0)
                out_bytes, out_name, issues = _process_workbook_bytes(content, f.name)
                report.extend(issues or [])
                if out_bytes is not None:
                    modified.append((out_name, out_bytes))
            except Exception as e:
                report.append(f"{f.name}: unexpected error: {e}")

        if not modified:
            st.warning("No files produced. Check the report below.")
        else:
            with st.expander("Download pre-processed files (optional)", expanded=False):
                if len(modified) == 1:
                    name, data = modified[0]
                    st.download_button(
                        f"Download {name}", data=data, file_name=name,
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        use_container_width=True
                    )
                else:
                    buf = io.BytesIO()
                    with zipfile.ZipFile(buf, "w", compression=zipfile.ZIP_DEFLATED) as zf:
                        for name, data in modified:
                            zf.writestr(name, data)
                    buf.seek(0)
                    st.download_button(
                        "Download all (ZIP)", data=buf, file_name="modified_excels.zip",
                        mime="application/zip", use_container_width=True
                    )

            if auto_continue:
                class TempUpload:
                    def __init__(self, data: bytes, name: str):
                        self._fh = tempfile.SpooledTemporaryFile(max_size=10_000_000, mode="w+b", suffix=".xlsx")
                        self._fh.write(data); self._fh.seek(0)
                        self.name = name
                    def read(self, *a, **k): return self._fh.read(*a, **k)
                    def seek(self, *a, **k): return self._fh.seek(*a, **k)
                    def tell(self, *a, **k): return self._fh.tell(*a, **k)
                    def readable(self): return True
                    def seekable(self): return True
                    def close(self): return self._fh.close()

                st.session_state["_forwarded_preproc_files"] = [TempUpload(b, n) for n, b in modified]
                st.success("Pre-processed files are forwarded. Open the main page and select Raw cycler Excel.")

        if report:
            st.subheader("Report")
            for line in report:
                st.write("-", line)
