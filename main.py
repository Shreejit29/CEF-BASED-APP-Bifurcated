import streamlit as st
import io
import pandas as pd
import numpy as np
from processing import (
    load_excel_first_sheet, build_final_dataset,
    _read_any, validate_processed_dataframe, ALIAS_MAP
)
from analysis import first_n_cef_stats
from viz import cef_figure, efficiencies_figure, capacities_figure
from utils import canonicalize_columns
import joblib, os
from train_cef_model import load_excel_features, train_from_dataframe, save_model

# NEW: imports for pre-processor
import zipfile
import openpyxl

st.set_page_config(page_title="Battery Health Prediction - CEF Analysis", page_icon="🔋", layout="wide")
st.title("🔋 Battery Health Prediction - CEF Analysis")
st.markdown("Upload raw cycler data for full processing or upload an already processed dataset to compute CEF statistics and plots")

# Target feature proportions
TARGET_W = {"cef_slope":0.364803, "cef_std":0.286279, "cef_range":0.216227, "cef_var":0.132692}

# ---------------- Sidebar ----------------
st.sidebar.header("Mode")
mode = st.sidebar.radio(
    "Choose input type",
    ("Pre-Processor (Excel)", "Raw cycler Excel", "Processed dataset"),
    help="Pre-Processor cleans Excel structure only; Raw performs full pipeline; Processed skips to analysis."
)

remove_first_row = st.sidebar.checkbox(
    "Remove First Row (Conditioning Cycle)", value=True,
    help="Applies only to raw cycler Excel processing"
)

SLOPE_WEIGHT = st.sidebar.number_input(
    "CEF slope weight", min_value=0.1, max_value=20.0, value=1.0, step=0.1,
    help="Multiply cef_slope by this factor before scaling. Keep consistent across training, prediction, and SHAP."
)
K_MULT = st.sidebar.number_input(
    "Target-importance strength k", min_value=0.0, max_value=20.0, value=5.0, step=0.5,
    help="Higher k pushes the model to emphasize the target feature proportions."
)

# -------- Persistent model load / upload --------
MODEL_PATH_DEFAULT = "models/cef_gb_model.joblib"
model = None
if os.path.exists(MODEL_PATH_DEFAULT):
    try:
        model = joblib.load(MODEL_PATH_DEFAULT)
        st.sidebar.success("Loaded saved model.")
    except Exception:
        st.sidebar.warning("Found saved model but failed to load.")

uploaded_model = st.sidebar.file_uploader("Upload trained model (.joblib)", type=["joblib"])
if uploaded_model is not None:
    try:
        model = joblib.load(uploaded_model)
        st.sidebar.success("Model loaded from upload.")
    except Exception as e:
        st.sidebar.error(f"Load failed: {e}")

# ---------------- Pre-Processor helpers (openpyxl) ----------------
COLUMNS_TO_DELETE = [3, 9, 10, 11, 12, 13]  # 1-based indices
NEW_HEADINGS = [
    "Cycle number",
    "Time",
    "Date",
    "Voltage (mV)",
    "Current (mA)",
    "Capacity (mAh)",
    "Energy (mWh)"
]

def _preproc_process_workbook(file_bytes: bytes, filename: str):
    issues = []
    try:
        wb = openpyxl.load_workbook(io.BytesIO(file_bytes))
    except Exception as e:
        return None, f"modified_{filename}", [f"{filename}: load error: {e}"]

    sheet_names = wb.sheetnames
    if len(sheet_names) < 2:
        return None, f"modified_{filename}", [f"{filename}: not enough sheets; skipped."]

    second_sheet_name = sheet_names[1]
    ws = wb[second_sheet_name]

    # delete other sheets
    for sn in list(sheet_names):
        if sn != second_sheet_name:
            wb.remove(wb[sn])

    # delete target columns in reverse order
    for col in sorted(COLUMNS_TO_DELETE, reverse=True):
        try:
            ws.delete_cols(col)
        except Exception as e:
            issues.append(f"{filename}: delete col {col} failed: {e}")

    # set new headers
    for idx, heading in enumerate(NEW_HEADINGS, start=1):
        ws.cell(row=1, column=idx, value=heading)

    # rename sheet to base filename
    base = filename.rsplit(".", 1)[0]
    ws.title = base

    # serialize
    out = io.BytesIO()
    wb.save(out)
    out.seek(0)
    return out.read(), f"modified_{filename}", issues

def compute_derivables(df: pd.DataFrame):
    if "Coulombic_Efficiency" not in df.columns and {"Discharge_Capacity","Charge_Capacity"}.issubset(df.columns):
        with np.errstate(divide='ignore', invalid='ignore'):
            df["Coulombic_Efficiency"] = np.where(df["Charge_Capacity"]>0,
                                                  df["Discharge_Capacity"]/df["Charge_Capacity"], np.nan)
    if "Energy_Efficiency" not in df.columns and {"Discharge_Energy","Charge_Energy"}.issubset(df.columns):
        with np.errstate(divide='ignore', invalid='ignore'):
            df["Energy_Efficiency"] = np.where(df["Charge_Energy"]>0,
                                               df["Discharge_Energy"]/df["Charge_Energy"], np.nan)
    if "CEF" not in df.columns and {"Coulombic_Efficiency","Energy_Efficiency"}.issubset(df.columns):
        CE = df["Coulombic_Efficiency"].astype(float)
        EE = df["Energy_Efficiency"].astype(float)
        df["CEF"] = 2 / (1 / np.exp(-10*(1-CE)) + 1 / np.exp(-10*(1-EE)))
    if "Cycle_Number" not in df.columns:
        df = df.reset_index(drop=True)
        df.insert(0, "Cycle_Number", range(1, len(df)+1))
    return df

def _features_from_stats(stats):
    slope = stats.get("slope"); rng = stats.get("range"); std = stats.get("std")
    var = (std ** 2) if std is not None else None
    if None in (slope, rng, std, var):
        return None
    arr = np.array([[slope, rng, std, var]], dtype=float)
    mul = np.array([
        (1.0 + K_MULT * TARGET_W["cef_slope"]) * float(SLOPE_WEIGHT),
        (1.0 + K_MULT * TARGET_W["cef_range"]),
        (1.0 + K_MULT * TARGET_W["cef_std"]),
        (1.0 + K_MULT * TARGET_W["cef_var"]),
    ], dtype=float)
    arr = arr * mul
    return arr

# ---------------- Page logic ----------------

# ... keep imports and helpers as in previous version ...

if mode == "Pre-Processor (Excel)":
    st.subheader("Pre-Processor: Excel structural cleanup")
    st.caption("Keeps only the 2nd sheet, deletes columns [3, 9, 10, 11, 12, 13], rewrites headers, and renames the sheet to the filename.")
    uploaded_files = st.file_uploader("Upload .xlsx files", type=["xlsx"], accept_multiple_files=True)

    # NEW control: continue into processing
    auto_continue = st.checkbox("Send pre-processed files to Raw Processing automatically", value=True)

    if st.button("Run Pre-Processor") and uploaded_files:
        modified = []
        report = []
        for uf in uploaded_files:
            try:
                content = uf.read(); uf.seek(0)
                out_bytes, out_name, issues = _preproc_process_workbook(content, uf.name)
                report.extend(issues or [])
                if out_bytes is not None:
                    modified.append((out_name, out_bytes))
            except Exception as e:
                report.append(f"{uf.name}: unexpected error: {e}")

        # Show report
        if report:
            st.markdown("#### Report")
            for line in report:
                st.write("-", line)

        if not modified:
            st.info("No files produced.")
            st.stop()

        # Offer download as before
        with st.expander("Download pre-processed files (optional)", expanded=False):
            if len(modified) == 1:
                name, data = modified[0]
                st.download_button(
                    f"Download {name}", data=data, file_name=name,
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )
            else:
                buf = io.BytesIO()
                import zipfile
                with zipfile.ZipFile(buf, "w", compression=zipfile.ZIP_DEFLATED) as zf:
                    for name, data in modified:
                        zf.writestr(name, data)
                buf.seek(0)
                st.download_button("Download all (ZIP)", data=buf, file_name="modified_excels.zip", mime="application/zip")

        # NEW: feed forward into Raw pipeline in-memory
        if auto_continue:
            st.success("Sending pre-processed files to Raw Processing…")

            # Emulate uploads with in-memory BytesIO objects that have .name attribute
            class InMemUpload(io.BytesIO):
                def __init__(self, data, name):
                    super().__init__(data)
                    self.name = name

            forwarded = [InMemUpload(b, n) for n, b in modified]

            # Flip the mode for the remainder of the script
            mode = "Raw cycler Excel"

            # Provide the forwarded list to the existing processing loop by setting a session state var
            st.session_state["_forwarded_preproc_files"] = forwarded

            # Continue to rest of script without st.stop()
        else:
            st.stop()

# ---------------- Existing flows ----------------

# If coming from Pre-Processor, use forwarded files instead of fresh uploader selection
if "_forwarded_preproc_files" in st.session_state:
    uploaded_files = st.session_state.pop("_forwarded_preproc_files")
else:
    uploaded_files = st.file_uploader(
        "Upload files",
        type=(['xlsx','xls'] if mode == "Raw cycler Excel" else ['csv','xlsx','xls']),
        accept_multiple_files=True
    )


# ---------------- Existing flows (unchanged) ----------------

uploaded_files = st.file_uploader(
    "Upload files",
    type=(['xlsx','xls'] if mode == "Raw cycler Excel" else ['csv','xlsx','xls']),
    accept_multiple_files=True
)

summary_rows = []

if uploaded_files is not None:
    for uploaded_file in uploaded_files:
        try:
            final_dataset = None
            st.write(f"### Processing file: {uploaded_file.name}")

            if mode == "Raw cycler Excel":
                sheet_name, df_raw = load_excel_first_sheet(uploaded_file)
                st.success(f"File uploaded successfully! Sheet: {sheet_name}")
                st.subheader("📊 Original Data Preview")
                st.write(f"Dataset shape: {df_raw.shape}")
                st.dataframe(df_raw.head())

                with st.expander(f"🧪 Data Preparation for {uploaded_file.name} (edit, rename, auto-map, and clean)", expanded=True):
                    st.caption("Edit cells, rename columns, and let the app recognize common aliases before processing.")
                    editable = st.data_editor(
                        df_raw, num_rows="dynamic", use_container_width=True,
                        key=f"data_editor_raw_{uploaded_file.name}"
                    )
                    df_for_next = editable.copy()
                    rename_map = {}
                    for col in editable.columns:
                        new = st.text_input(f"Rename '{col}' to:", value=col, key=f"rename_raw_{col}_{uploaded_file.name}")
                        if new and new != col:
                            rename_map[col] = new
                    if rename_map:
                        df_for_next = df_for_next.rename(columns=rename_map)

                    df_for_next, canon_notes = canonicalize_columns(df_for_next, ALIAS_MAP)
                    if canon_notes:
                        st.info(" | ".join(canon_notes))

                    st.write("Preview after renaming & canonicalization:")
                    st.dataframe(df_for_next.head())

                    required = ["Time","Date","Current (mA)","Capacity (mAh)","Energy (mWh)"]
                    can_process = all(k in df_for_next.columns for k in required)
                    if not can_process:
                        missing = [k for k in required if k not in df_for_next.columns]
                        st.error(f"Missing required columns for raw processing: {missing}")
                    else:
                        with st.spinner("Processing raw data..."):
                            final_dataset = build_final_dataset(df_for_next, remove_first_row)

            else:
                df_loaded, fmt = _read_any(uploaded_file)
                st.success(f"Processed dataset loaded ({fmt}).")

                with st.expander(f"🧪 Data Preparation for {uploaded_file.name} (edit, rename, auto-map, compute missing fields)", expanded=True):
                    editable = st.data_editor(
                        df_loaded, num_rows="dynamic", use_container_width=True,
                        key=f"data_editor_proc_{uploaded_file.name}"
                    )
                    df_for_next = editable.copy()
                    rename_map = {}
                    for col in editable.columns:
                        new = st.text_input(f"Rename '{col}' to:", value=col, key=f"rename_proc_{col}_{uploaded_file.name}")
                        if new and new != col:
                            rename_map[col] = new
                    if rename_map:
                        df_for_next = df_for_next.rename(columns=rename_map)

                    df_for_next, canon_notes = canonicalize_columns(df_for_next, ALIAS_MAP)
                    if canon_notes:
                        st.info(" | ".join(canon_notes))

                    df_ready = compute_derivables(df_for_next)
                    try:
                        final_dataset, notes = validate_processed_dataframe(df_ready)
                    except Exception as e:
                        final_dataset = df_ready
                        notes = [f"Validation fallback: {e}"]
                    if notes:
                        st.warning("Notes: " + " | ".join(notes))

            if final_dataset is None or final_dataset.empty:
                st.warning("No valid rows available for analysis after preparation. Please adjust the data and try again.")
            else:
                st.subheader(f"🔧 Data Preview For Analysis: {uploaded_file.name}")
                st.write(f"Final dataset shape: {final_dataset.shape}")
                st.dataframe(final_dataset.head(10))

                missing_eff = [c for c in ["Coulombic_Efficiency","Energy_Efficiency"] if c not in final_dataset.columns]
                if missing_eff:
                    st.info(f"Efficiency columns missing: {missing_eff}. They are computed when capacity/energy pairs are present.")

                st.subheader(f"📈 CEF Analysis - First 10 Cycles: {uploaded_file.name}")
                stats = first_n_cef_stats(final_dataset, first_n=10)

                cef_var = None
                if stats.get("std") is not None:
                    cef_var = float(stats["std"] ** 2)
                else:
                    try:
                        if "CEF" in stats["first_n_df"].columns and len(stats["first_n_df"]["CEF"]) >= 2:
                            cef_var = float(np.var(stats["first_n_df"]["CEF"].to_numpy(dtype=float), ddof=1))
                    except Exception:
                        cef_var = None

                col1, col2, col3, col4 = st.columns(4)
                col1.metric("CEF Slope", f"{stats['slope']:.6f}" if stats['slope'] is not None else "N/A", help="Linear regression slope")
                col2.metric("CEF Range", f"{stats['range']:.6f}" if stats['range'] is not None else "N/A", help="Max - Min CEF values")
                col3.metric("CEF Std Dev", f"{stats['std']:.6f}" if stats['std'] is not None else "N/A", help="Sample standard deviation (ddof=1)")
                col4.metric("CEF Variance", f"{cef_var:.6f}" if cef_var is not None else "N/A", help="Sample variance (ddof=1)")

                st.plotly_chart(cef_figure(stats), use_container_width=True)

                st.subheader(f"📊 Additional Analysis: {uploaded_file.name}")
                c1, c2 = st.columns(2)
                with c1:
                    st.plotly_chart(efficiencies_figure(stats), use_container_width=True)
                with c2:
                    st.plotly_chart(capacities_figure(stats), use_container_width=True)

                st.subheader(f"🧪 Condition prediction: {uploaded_file.name}")
                if model is None:
                    st.info("Load or train a model to enable predictions.")
                else:
                    feats = _features_from_stats(stats)
                    if feats is None:
                        st.info("Not enough data to form features for inference (need slope, range, std).")
                    else:
                        try:
                            proba = getattr(model, "predict_proba", None)
                            if proba is not None and hasattr(model, "classes_"):
                                probs = proba(feats)[0]
                                if 1 in model.classes_:
                                    pos_idx = list(model.classes_).index(1)
                                else:
                                    pos_idx = 1 if len(model.classes_) > 1 else 0
                                p = float(probs[pos_idx])
                                pred = int(p >= 0.5)
                            else:
                                pred = int(model.predict(feats)[0])
                                p = None
                            label = "Cell will degrade in future" if pred == 1 else "Cell will remain healthy in future"
                            if p is not None:
                                st.metric("Predicted condition", f"{label}", f"Degraded prob: {p:.2%}")
                            else:
                                st.metric("Predicted condition", f"{label}")
                        except Exception as e:
                            st.warning(f"Inference error: {e}")

                st.subheader(f"💾 Download Results: {uploaded_file.name}")
                statistics_df = pd.DataFrame({
                    'Parameter': ['CEF Slope (Linear Regression)', 'CEF Range', 'CEF Standard Deviation', 'CEF Variance'],
                    'Value': [stats['slope'], stats['range'], stats['std'], cef_var],
                    'Description': [
                        'Linear regression slope of CEF vs Cycle Number for first 10 cycles',
                        'Difference between maximum and minimum CEF values in first 10 cycles',
                        'Sample standard deviation (ddof=1) of CEF for first 10 cycles',
                        'Sample variance (ddof=1) of CEF for first 10 cycles'
                    ]
                })
                output = io.BytesIO()
                with pd.ExcelWriter(output, engine='xlsxwriter') as writer:
                    statistics_df.to_excel(writer, sheet_name='CEF_Statistics', index=False)
                    stats['first_n_df'].to_excel(writer, sheet_name='First_10_Cycles', index=False)
                    final_dataset.to_excel(writer, sheet_name='Complete_Dataset', index=False)
                st.download_button(
                    "📥 Download Complete Analysis (Excel)",
                    output.getvalue(),
                    file_name=f"CEF_Analysis_Results_{uploaded_file.name}",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    key=f"download_excel_{uploaded_file.name}"
                )
                st.download_button(
                    "📄 Download Dataset (CSV)",
                    final_dataset.to_csv(index=False),
                    file_name=f"processed_battery_data_{uploaded_file.name}.csv",
                    mime="text/csv",
                    key=f"download_csv_{uploaded_file.name}"
                )

                summary_rows.append({
                    "FileName": uploaded_file.name,
                    "CEF_Slope": stats.get("slope"),
                    "CEF_Range": stats.get("range"),
                    "CEF_StdDev": stats.get("std"),
                    "CEF_Variance": cef_var,
                    "Prediction": label,
                    "Degraded_Probability": p
                })

        except Exception as e:
            st.error(f"Error: {str(e)}")

    if summary_rows:
        summary_df = pd.DataFrame(summary_rows)
        st.subheader("📋 Summary of Predictions")
        st.dataframe(summary_df)

        csv_summary = summary_df.to_csv(index=False).encode()
        excel_buffer = io.BytesIO()
        with pd.ExcelWriter(excel_buffer, engine='xlsxwriter') as writer:
            summary_df.to_excel(writer, sheet_name="Summary", index=False)
        excel_data = excel_buffer.getvalue()

        st.download_button(
            "Download Summary CSV",
            csv_summary,
            file_name="battery_health_summary.csv",
            mime="text/csv",
            key="download_summary_csv"
        )
        st.download_button(
            "Download Summary Excel",
            excel_data,
            file_name="battery_health_summary.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            key="download_summary_excel"
        )
else:
    st.info("👆 Upload a file to begin analysis")

st.subheader("🧠 Train classifier (Excel)")
with st.expander("Train Gradient Boosting on labeled CEF stats", expanded=False):
    st.caption("Excel must have: cell_id, cef_slope, cef_range, cef_std, optional cef_var, label (0=Healthy, 1=Degraded).")
    excel_file = st.file_uploader("Upload labeled Excel", type=["xlsx","xls"], key="train_xlsx")
    if excel_file is not None:
        try:
            df_train = load_excel_features(excel_file, sheet=0)
            st.write(f"Rows loaded: {len(df_train)}")
            st.dataframe(df_train.head())

            if st.button("Train model", type="primary"):
                with st.spinner("Training Gradient Boosting (grouped CV)…"):
                    res = train_from_dataframe(df_train, random_state=42, slope_weight=SLOPE_WEIGHT, k=K_MULT)
                st.success("Training complete")
                st.write("Model parameters:", res["best_params"])
                if res["cv_f1_mean"] is not None:
                    st.write(f"CV F1 score: {res['cv_f1_mean']:.4f} ± {res['cv_f1_std']:.4f}")
                st.text("Test set classification report:\n" + res["test_report"])
                if res["test_auc"] is not None:
                    st.write(f"Test AUC: {res['test_auc']:.4f}")
                st.write("Confusion matrix:", res["test_confusion_matrix"])

                os.makedirs("models", exist_ok=True)
                save_path = save_model(res["model"], path="models/cef_gb_model.joblib")
                st.success(f"Model saved to {save_path} and will auto-load next time.")
                model = res["model"]  # activate immediately
        except Exception as e:
            st.error(f"Training error: {e}")

with st.expander("🔍 Model explainability (SHAP)", expanded=False):
    st.caption("Compute and display SHAP feature importance for the current trained model using the labeled CEF stats Excel.")
    shap_excel = st.file_uploader("Upload labeled CEF stats Excel for SHAP", type=["xlsx","xls"], key="shap_xlsx")
    if shap_excel is not None and model is not None:
        try:
            df_shap = load_excel_features(shap_excel, sheet=0)
            feature_names = ["cef_slope","cef_range","cef_std","cef_var"]
            X_all = df_shap[feature_names].values.astype(float)
            mul = np.array([
                (1.0 + K_MULT * TARGET_W["cef_slope"]) * float(SLOPE_WEIGHT),
                (1.0 + K_MULT * TARGET_W["cef_range"]),
                (1.0 + K_MULT * TARGET_W["cef_std"]),
                (1.0 + K_MULT * TARGET_W["cef_var"]),
            ], dtype=float)
            X_all = X_all * mul

            scaler = getattr(model, "named_steps", {}).get("scaler", None)
            clf = getattr(model, "named_steps", {}).get("gbc", None)
            if clf is None:
                st.info("Current model is not a GradientBoosting pipeline.")
            else:
                X_scaled = scaler.transform(X_all) if scaler is not None else X_all
                import shap, matplotlib.pyplot as plt
                explainer = shap.TreeExplainer(clf)
                sv = explainer.shap_values(X_scaled)
                shap_values = sv[1] if isinstance(sv, list) and len(sv)>=2 else sv

                mean_abs = np.mean(np.abs(shap_values), axis=0)
                shap_df = pd.DataFrame({"feature": feature_names, "mean_abs_shap": mean_abs}).sort_values("mean_abs_shap", ascending=False)
                st.dataframe(shap_df)

                fig1 = plt.figure()
                shap.summary_plot(shap_values, features=X_scaled, feature_names=feature_names, show=False)
                st.pyplot(fig1); plt.close(fig1)

                fig2 = plt.figure()
                shap.summary_plot(shap_values, features=X_scaled, feature_names=feature_names, plot_type="bar", show=False)
                st.pyplot(fig2); plt.close(fig2)
        except Exception as e:
            st.error(f"SHAP error: {e}")
    elif model is None:
        st.info("Train or load a model first.")
